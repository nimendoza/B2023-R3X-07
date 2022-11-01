# These imports are needed to have this segment of the software run
from openpyxl                      import load_workbook
from openpyxl.utils                import get_column_letter
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet  import Worksheet
from os      import walk
from os.path import exists
from xlsxwriter import Workbook

# Read and write agent implementation in Python 3.10.6
class ReadWriteAgent:
  # Generate a .xlsx file
  @classmethod
  def generate_xlsx(cls, path: str):
    workbook = Workbook(path)
    workbook.close()

  # Return the literal, integer, or string representation of a value
  @classmethod
  def as_text(cls, value):
    if value is None:
      return value
    if any([
      isinstance(value, float),
      isinstance(value, int),
      isinstance(value, str) and value.isdigit()
    ]):
      return int(value)
    return str(value)
    
  # All newly generated .xlsx files have a "Sheet1" worksheet. This 
  # function deletes it
  @classmethod
  def delete_default_sheet(cls, path: str):
    workbook = load_workbook(path, read_only=False)
    if 'Sheet1' in workbook.sheetnames:
      worksheet = workbook['Sheet1']
      if not isinstance(worksheet, ReadOnlyWorksheet):
        workbook.remove(worksheet)
    workbook.save(path)
    workbook.close()

  # Make columns "fit" the value it contains
  @classmethod
  def realign_columns(cls, path: str, sheet: str):
    workbook = load_workbook(path)
    worksheet = workbook[sheet]
    if isinstance(worksheet, Worksheet):
      for column_cells in worksheet.columns:
        worksheet.column_dimensions[
          get_column_letter(column_cells[0].column)
        ].width = max(len(repr(cell.value)) for cell in column_cells)
    workbook.save(path)
    workbook.close()
      
  # Given a worksheet of a workbook, return a 2D array of its data
  @classmethod
  def read_xlsx(cls, path: str, sheet: str):
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet]
    data = list[list]()
    if isinstance(worksheet, Worksheet):
      for row in worksheet.rows:
        data.append(list(
          ReadWriteAgent.as_text(cell.value) 
            for cell in row
        ))
    workbook.close()
    return data
    
  # Given a 2D array, write it as a worksheet of the workbook indicated
  # in "path"
  @classmethod
  def to_xlsx(cls, path: str, sheet: str, data: list[list]):
    if not exists(path):
      ReadWriteAgent.generate_xlsx(path)
      
    workbook = load_workbook(path)
    worksheet = workbook.create_sheet(sheet)
    for row in data:
      worksheet.append(ReadWriteAgent.as_text(cell) for cell in row)
    workbook.save(path)
    workbook.close()
    
    ReadWriteAgent.delete_default_sheet(path)
    ReadWriteAgent.realign_columns(path, sheet)
    
  # Given a name template and a directory, return an indexed string
  # of that name template that hasn't been "taken" yet
  @classmethod
  def find_filepath(cls, directory: str, name_template: str) -> str:
    path_template = f'{directory}/{name_template}'
    for index in range(len(next(walk(directory))[2]) + 1):
      if not exists(path_template.format(index)):
        return path_template.format(index)
    return path_template.format(0)
      
  # Given an .xlsx file, return the worksheets of that workbook
  @classmethod
  def get_sheetnames(cls, path: str):
    workbook = load_workbook(path)
    sheetnames = set(workbook.sheetnames)
    workbook.close()
    return sheetnames