from __future__  import annotations
from collections import defaultdict
from copy        import deepcopy
from datetime    import date
from math        import ceil
from openpyxl                      import load_workbook
from openpyxl.utils                import get_column_letter
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet  import Worksheet
from os         import walk
from os.path    import exists
from time       import time
from typing     import Any, Iterable, Optional, Union
from xlsxwriter import Workbook

import random
import sys
import traceback

class Session:
  alias: str
  
  def __init__(self, alias: str):
    self.alias = alias
    
  def __repr__(self):
    return self.alias
  
  def __lt__(self, other: Session):
    return self.alias < other.alias

class Shift:
  sessions: list[Session]
  
  def __init__(self):
    self.sessions = list()
    
  def __repr__(self):
    return ''.join(map(str, self.sessions))
  
  def __lt__(self, other: Shift):
    return str(self) < str(other)
  
  def add(self, session: Session):
    self.sessions.append(session)
    self.sessions.sort()
    
class Capacity:
  minimum: int
  ideal  : int
  maximum: int
  
  def __init__(self, minimum: int, ideal: int, maximum: int):
    self.minimum = minimum
    self.ideal   = ideal
    self.maximum = maximum
    
class CourseType:
  alias: str
  order: int
  
  def __init__(self, alias: str, order: int):
    self.alias = alias
    self.order = order
    
  def __repr__(self):
    return self.alias
  
  def __lt__(self, other: CourseType):
    return self.order < other.order
  
class ParallelSession:
  shift  : Shift
  session: Optional[Session]
  index  : Optional[int]
  
  def __init__(
    self,
    shift  : Shift,
    session: Optional[Session] = None,
    index  : Optional[int]     = None):
    self.shift   = shift
    self.session = session
    self.index   = index
    
  def __repr__(self):
    return f'{self.session or self.shift}{self.index or ""}'
  
  def __contains__(self, value: Union[Session, Shift]):
    return value in {self.session, self.shift}
  
class GradeLevel:
  alias  : int
  courses: dict[tuple[CourseType, bool], set[Course]]
  
  def __init__(self, alias: int):
    self.alias   = alias
    self.courses = defaultdict(set)
    
  def __repr__(self):
    return f'Grade {self.alias}'
  
  def __lt__(self, other: GradeLevel):
    return str(self) < str(other)
  
  def add(self, course_type: CourseType, ranked: bool, course: Course):
    self.courses[(course_type, ranked)].add(course)
    
class Ranking:
  ordered_courses: dict[CourseType, list[Course]]
  reason_rejected: dict[CourseType, dict[Course, str]]
  
  def __init__(self, grade_level: GradeLevel):
    self.ordered_courses = dict(
      (course_type, list())
      for course_type, ranked in grade_level.courses if ranked)
    self.reason_rejected = dict(
      (course_type, dict())
      for course_type, ranked in grade_level.courses if ranked)
    
  def len(self, course_type: CourseType):
    return len(self.ordered_courses[course_type])
  
  def add(self, course_type: CourseType, course: Course):
    self.ordered_courses[course_type].append(course)
    
  def pop(self, course_type: CourseType, reason: str, index: int):
    course = self.ordered_courses[course_type].pop(index)
    self.reason_rejected[course_type][course] = reason
    
class Rankings:
  owner: Student
  start: Ranking
  final: Ranking
  
  def __init__(self, owner: Student):
    self.owner = owner
    self.start = Ranking(self.owner.grade_level)
    self.final = Ranking(self.owner.grade_level)
    
  def add(self, course_type: CourseType, course: Course):
    if course.qualified(self.owner):
      self.start.add(course_type, course)
      self.final.add(course_type, course)
      
  def initial(self, course_type: CourseType, index: Optional[int] = None):
    if not self.start.len(course_type):
      return None
    return self.start.ordered_courses[course_type][index or 0]
  
  def current(self, course_type: CourseType, index: Optional[int] = None):
    if not self.final.len(course_type):
      for course in self.owner.grade_level.courses[(course_type, True)]:
        if course.qualified(self.owner):
          self.final.add(course_type, course)
      self.final.ordered_courses[course_type].sort()
    return self.final.ordered_courses[course_type][index or 0]
  
  def reset(self):
    for course_type in self.start.ordered_courses:
      self.final.ordered_courses[course_type] = list(
        self.start.ordered_courses[course_type])
  
class ResearchGroup:
  course  : Course
  alias   : str
  students: set[Student]
  __shift : Optional[Shift]
  
  def __init__(self, course: Course, alias: str):
    self.alias    = alias
    self.course   = course
    self.students = set()
    self.__shift  = None
    
  def __repr__(self):
    return f'{self.course} {self.alias}'
  
  def __lt__(self, other: ResearchGroup):
    return str(self) < str(other)
  
  @property
  def shift(self):
    return self.__shift
  
  @shift.setter
  def shift(self, value: Optional[Shift]):
    self.__shift = value
    for student in self.students:
      if student.shift != value:
        student.shift = value
        
  def add(self, student: Student):
    student.research_group = self
    self.students.add(student)
    
class Student:
  grade_level: GradeLevel
  alias      : str
  __shift    : Optional[Shift]
  
  rankings: Rankings
  taken   : set[Course]
  takes   : dict[CourseType, Course]
  
  research_group: Optional[ResearchGroup]
  sessions      : set[Session]
  sections      : dict[Course, Section]
  
  def __init__(self, alias: str, grade_level: GradeLevel):
    self.alias       = alias
    self.grade_level = grade_level
    self.__shift     = None
    
    self.rankings = Rankings(self)
    self.taken    = set()
    self.takes    = dict()
    
    self.research_group = None
    self.sessions       = set()
    self.sections       = dict()
    
  def __repr__(self):
    return f'{self.grade_level}-{self.alias}'
  
  def __lt__(self, other: Student):
    return str(self) < str(other)
  
  @property
  def shift(self):
    return self.__shift
  
  @shift.setter
  def shift(self, value: Optional[Shift]):
    self.__shift = value
    if self.research_group and self.research_group.shift != value:
      self.research_group.shift = value
      
  @property
  def available_sessions(self):
    sessions = set[Session]()
    if self.shift:
      sessions.update(self.shift.sessions)
      sessions.difference_update(self.sessions)
    return sorted(sessions)
  
  @property
  def has_taken_level_two(self):
    for course in self.taken:
      if course.difficulty_level == 2:
        return True
    return False
  
class Section:
  course          : Course
  parallel_session: ParallelSession
  capacity        : Capacity
  students        : set[Student]
  
  def __init__(self, course: Course, parallel_session: ParallelSession):
    self.course           = course
    self.parallel_session = parallel_session
    self.capacity         = self.course.capacity_section
    self.students         = set()
    
  def __repr__(self):
    return '{} {}'.format(self.course, self.parallel_session)
  
  def __lt__(self, other: Section):
    return str(self) < str(other)
  
  def qualified(self, student: Student):
    return all([
      self.course.qualified(student),
      student.shift in {self.parallel_session.shift, None},
      self.parallel_session.session in student.available_sessions
      if self.parallel_session.session and student.shift else True])
    
  def add(self, student: Student, course_type: CourseType):
    if self.qualified(student):
      self.students.add(student)
      student.shift = self.parallel_session.shift
      student.takes[course_type] = self.course
      student.sections[self.course] = self
      if self.parallel_session.session:
        student.sessions.add(self.parallel_session.session)
      return True
    return False
  
  def overload(self, student: Student, course_type: CourseType):
    if len(self.students) < self.capacity.maximum:
      return self.add(student, course_type)
    return False
  
  def enroll(self, student: Student, course_type: CourseType):
    if len(self.students) < self.capacity.ideal:
      return self.add(student, course_type)
    return False
  
class Course:
  alias           : str
  difficulty_level: int
  linked_to       : Optional[Course]
  
  capacity_section : Capacity
  capacity_sections: Capacity
  sections         : list[Section]
  
  not_alongside: set[Course]
  prerequisites: list[set[Course]]
  
  def __init__(self, alias: str, difficulty_level: int):
    self.alias            = alias
    self.difficulty_level = difficulty_level
    self.linked_to        = None
    
    self.sections      = list()
    self.not_alongside = {self}
    self.prerequisites = list()
    
  def __repr__(self):
    return '{}{}'.format(
      self.alias,
      f' Level {self.difficulty_level}' if self.difficulty_level else '')
    
  def __lt__(self, other: Course):
    return str(self) < str(other)
  
  @property
  def could_open_section(self):
    return sum(
      len(course.sections) for course in [self, self.linked_to] if course
    ) < self.capacity_sections.maximum
    
  def list_sections_by(self, value: Union[Session, Shift]):
    result = list[Section]()
    for section in self.sections:
      if value in section.parallel_session:
        result.append(section)
    return sorted(result, key=lambda section: len(section.students))
  
  def qualified(self, student: Student):
    return all([
      self not in student.taken,
      self.not_alongside.isdisjoint(student.takes),
      all(
        prerequisites.intersection(student.taken)
        for prerequisites in self.prerequisites)])
    
  def overload(self, student: Student, course_type: CourseType):
    try:
      return min(filter(
        lambda section: section.qualified(student),
        self.sections), key=lambda section: len(section.students)
      ).overload(student, course_type)
    except ValueError:
      return False
    
  def enroll(self, student: Student, course_type: CourseType):
    try:
      return min(filter(
        lambda section: section.qualified(student),
        self.sections), key=lambda section: len(section.students)
      ).enroll(student, course_type)
    except ValueError:
      return False
  
class Data:
  def __init__(self):
    self.shifts          = set[Shift]()
    self.course_types    = dict[str, CourseType]()
    self.courses         = dict[str, Course]()
    self.grade_levels    = dict[str, GradeLevel]()
    self.research_groups = dict[str, ResearchGroup]()
    self.students        = defaultdict[GradeLevel, list[Student]](list)

  def reset(self):
    for students in self.students.values():
      for student in students:
        for course_type in self.course_types.values():
          course = None
          if course_type in student.takes:
            course = student.takes.pop(course_type)
          if course in student.sections:
            student.sections.pop(course)
        student.sessions.clear()
        student.rankings.reset()
        student.shift = None
    for course in self.courses.values():
      course.sections.clear()

class SolutionV1:
  shifts      : list[Shift]
  courses     : list[Course]
  course_types: dict[str, CourseType]
  
  __research_groups : list[ResearchGroup]
  __students        : list[Student]
  __grouped_students: list[Student]
  __nogroup_students: list[Student]
  
  core: CourseType
  elec: CourseType
  math: CourseType
  res : CourseType
  
  def __init__(self, data: Data):
    self.course_types = data.course_types
    self.shifts       = sorted(data.shifts)
    self.courses      = sorted(data.courses.values())
    
    self.core = self.course_types[CORE]
    self.elec = self.course_types[ELEC]
    self.math = self.course_types[MATH]
    self.res  = self.course_types[RESEARCH]
    
    self.__research_groups  = list()
    self.__students         = list()
    self.__grouped_students = list()
    self.__nogroup_students = list()
    for student in data.students[data.grade_levels['Grade 12']]:
      if not student.has_taken_level_two:
        core = student.rankings.initial(self.core)
        if not core or core.difficulty_level != 2:
          while student.rankings.current(self.elec).difficulty_level != 2:
            student.rankings.final.pop(
              self.elec, 'Needs to take a level 2 course', 0)
        if not student.research_group:
          research = random.choice(list(
            student.grade_level.courses[(self.res, False)]))
          student.research_group = ResearchGroup(research, 'Temporary')
          student.research_group.add(student)
          
    research_groups  = set[ResearchGroup]()
    grouped_students = set[Student]()
    nogroup_students = set[Student]()
    for students in data.students.values():
      self.__students.extend(students)
      for student in students:
        if student.research_group:
          research_groups.add(student.research_group)
          grouped_students.add(student)
        else:
          nogroup_students.add(student)
    self.__students.sort()
    self.__research_groups  = sorted(research_groups)
    self.__grouped_students = sorted(grouped_students)
    self.__nogroup_students = sorted(nogroup_students)
    
  @property
  def students(self):
    random.shuffle(self.__students)
    return self.__students
  
  @property
  def research_groups(self):
    random.shuffle(self.__research_groups)
    return self.__research_groups
  
  @property
  def grouped_students(self):
    random.shuffle(self.__grouped_students)
    return self.__grouped_students
  
  @property
  def nogroup_students(self):
    random.shuffle(self.__nogroup_students)
    return self.__nogroup_students
  
  def open_sections_to_overload(self, course_type: CourseType):
    courses = set[Course](
      student.rankings.current(course_type) for student in self.students)
    for course in courses:
      for shift in self.shifts:
        course.sections.append(Section(course, ParallelSession(shift)))
        
  def open_sections_spread_out(self, course_type: CourseType):
    courses = set[Course](random.choice(list(student.grade_level.courses[(
      course_type, False)])) for student in self.students)
    for course in courses:
      for shift in self.shifts:
        for session in shift.sessions:
          course.sections.append(Section(course, ParallelSession(
            shift, session)))
      leftover = course.capacity_sections.maximum - len(course.sections)
      for _ in range(leftover):
        shift   = random.choice(self.shifts)
        session = random.choice(shift.sessions)
        course.sections.append(Section(course, ParallelSession(
          shift, session, len(course.list_sections_by(session)))))
        
  def get_course_demand(self, students: Iterable[Student]):
    demand = defaultdict[Course, int](int)
    for student in students:
      core = student.rankings.current(self.core)
      elec = student.rankings.current(self.elec)
      while elec in core.not_alongside:
        student.rankings.final.pop(self.elec, 'Not compatible with CSE', 0)
        elec = student.rankings.current(self.elec)
      demand[core] += 1
      demand[elec] += 1
    return demand
      
  def get_section_combinations(
    self, students: Iterable[Student], demand: dict[Course, int]):
    section_combinations = defaultdict[Shift, set[tuple[
      Student, Section, Section]]](set)
    for student in students:
      core = student.rankings.current(self.core)
      elec = student.rankings.current(self.elec)
      for csection in filter(
        lambda s: len(s.students) + demand[core] <= s.capacity.maximum,
        core.sections):
        for esection in filter(
          lambda s: len(s.students) + demand[core] <= s.capacity.maximum,
          elec.sections):
          if all([
            csection.parallel_session.shift 
            == esection.parallel_session.shift,
            csection.parallel_session.session 
            != esection.parallel_session.session]):
            section_combinations[csection.parallel_session.shift].add((
              student, csection, esection))
    
    paired_combinations = defaultdict[tuple[Shift, Session], list[tuple[
      Student, Section, Section]]](list)
    for shift in section_combinations:
      for session in shift.sessions:
        for student, csection, esection in section_combinations[shift]:
          if all([
            csection.parallel_session.session != session,
            esection.parallel_session.session != session]):
            paired_combinations[(shift, session)].append((
              student, csection, esection))
    return paired_combinations
  
  def section_student(self, student: Student, course_type: CourseType):
    course = student.rankings.current(course_type)
    if course.qualified(student):
      for session in student.available_sessions:
        for section in course.list_sections_by(session):
          if section.overload(student, course_type):
            return
      if course.could_open_section:
        assert student.shift
        session = random.choice(student.available_sessions)
        section = Section(course, ParallelSession(
          student.shift, session, len(course.list_sections_by(session))))
        course.sections.append(section)
        if not section.overload(student, course_type):
          raise Exception('Impossible')
        return
      student.rankings.final.pop(course_type, 'No rooms available', 0)
    else:
      student.rankings.final.pop(course_type, 'No qualified', 0)
    self.section_student(student, course_type)
  
  def section_grouped(
    self, 
    research_course: Course,
    students       : set[Student], 
    combinations   : dict[tuple[Shift, Session], list[tuple[
      Student, Section, Section]]]):
    research_sections = list(filter(
      lambda s: len(s.students) + len(students) <= s.capacity.ideal if any(
                len(g.students) < g.capacity.ideal
                for g in research_course.sections) else True,
      research_course.sections))
    if combinations:
      for rsection in research_sections:
        shift   = rsection.parallel_session.shift
        session = rsection.parallel_session.session
        if (shift, session) not in combinations:
          continue
        for student in students:
          rsection.overload(student, self.res)
          student.rankings.current(self.math).overload(student, self.math)

        assert session
        random.shuffle(combinations[(shift, session)])
        for student, csection, esection in combinations[(shift, session)]:
          if not set(student.takes).isdisjoint([self.core, self.elec]):
            csection.overload(student, self.core)
            esection.overload(student, self.elec)
        for student in students:
          for type in {self.core, self.elec}.difference(student.takes):
            self.section_student(student, type)
        return
    
    rsection = random.choice(research_sections)
    for student in students:
      rsection.overload(student, self.res)
      student.rankings.current(self.math).overload(student, self.math)
      for course_type in [self.core, self.elec]:
        self.section_student(student, course_type)
        
  def enroll_initial(self, student: Student):
    research = random.choice(list(student.grade_level.courses[(
      self.res, False)]))
    core = student.rankings.current(self.core)
    elec = student.rankings.current(self.elec)
    while elec in core.not_alongside:
      student.rankings.final.pop(self.elec, 'Not compatible with CSE', 0)
      elec = student.rankings.current(self.elec)
    for shift in self.shifts:
      for c in core.list_sections_by(shift):
        for e in elec.list_sections_by(shift):
          for r in research.list_sections_by(shift):
            if all([
              c.parallel_session.session != e.parallel_session.session,
              c.parallel_session.session != r.parallel_session.session,
              e.parallel_session.session != r.parallel_session.session,
              len(c.students) < c.capacity.maximum,
              len(e.students) < e.capacity.maximum,
              len(r.students) < r.capacity.ideal
              if any(
                len(s.students) < s.capacity.ideal
                for s in research.sections) else True]):
              if not all([
                c.overload(student, self.core),
                e.overload(student, self.elec),
                r.overload(student, self.res)]):
                raise Exception('Impossible')
              student.rankings.current(self.math).overload(
                student, self.math)
              return True
    return False
  
  def cleanup_student_rankings(self, student: Student):
    if self.elec not in student.takes:
      core = None
      if self.core in student.takes:
        core = student.takes[self.core]
      else:
        core = student.rankings.current(self.core)
        
      elec = student.rankings.current(self.elec)
      while elec in core.not_alongside:
        student.rankings.final.pop(self.elec, 'Not compatible with CSE', 0)
        elec = student.rankings.current(self.elec)
    if self.core not in student.takes:
      if self.elec in student.takes:
        elec      = student.takes[self.elec]
        core      = student.rankings.current(self.core)
        countdown = len(student.grade_level.courses[(self.core, True)]) + 1
        while core in elec.not_alongside and countdown:
          student.rankings.final.pop(
            self.core, 'Not compatible with STE', 0)
          core       = student.rankings.current(self.core)
          countdown -= 1
        if not countdown:
          raise Exception('Impossible')
  
  def open_sections_based_on(
    self, demand: defaultdict[Course, set[tuple[CourseType, Student]]]):
    good = False
    for course, pairs in demand.items():
      sessioned = defaultdict[
        tuple[Shift, Session], set[tuple[CourseType, Student]]](set)
      for course_type, student in pairs:
        assert student.shift
        for session in student.available_sessions:
          sessioned[(student.shift, session)].add((course_type, student))
      shift, session = max(sessioned, key=lambda x: len(sessioned[x]))
      if len(sessioned[(
        shift, session)]) >= course.capacity_section.minimum:
        if course.could_open_section:
          course.sections.append(Section(course, ParallelSession(
            shift, session, len(course.list_sections_by(session)))))
          for course_type, student in pairs:
            course.enroll(student, course_type)
          good = True
        else:
          for course_type, student in pairs:
            student.rankings.final.pop(
              course_type, 'No rooms available', 0)
            self.cleanup_student_rankings(student)
      else:
        for course_type, student in pairs:
          student.rankings.final.pop(
            course_type, 'Too few demand to open another room', 0)
          self.cleanup_student_rankings(student)
    return good
    
  def enroll_final(self, student: Student):
    for core in filter(
      lambda course: course.qualified(student),
      student.grade_level.courses[(self.core, True)]):
      for elec in filter(
        lambda course: all([
          course not in core.not_alongside,
          course.qualified(student)]),
        student.grade_level.courses[(self.elec, True)]):
        for c in core.sections:
          for e in elec.sections:
            if all([
              c.parallel_session.session in student.available_sessions,
              e.parallel_session.session in student.available_sessions,
              c.parallel_session.session != e.parallel_session.session,
              len(c.students) < c.capacity.maximum,
              len(e.students) < e.capacity.maximum]):
              if not c.overload(student, self.core):
                raise Exception('Impossible')
              if not e.overload(student, self.elec):
                raise Exception('Impossible')
              if core != student.rankings.current(self.core):
                student.rankings.final.pop(
                  self.core, 'Last resort sectioning', 0)
              if elec != student.rankings.current(self.elec):
                student.rankings.final.pop(
                  self.elec, 'Last resort sectioning', 0)
              return True
    return False
          
  def rebalance_sections(self):
    for course in set[Course](
      student.takes[self.course_types[course_type]]
      for student in self.students for course_type in [CORE, ELEC]):
      for shift in self.shifts:
        for session in shift.sessions:
          sections = course.list_sections_by(session)
          students = set[Student]()
          for section in sections:
            for student in list(section.students):
              section.students.remove(student)
              for type, course_ in list(student.takes.items()):
                if course_ == course:
                  student.sections.pop(student.takes.pop(type))
                  student.sessions.remove(
                    section.parallel_session.session)  # type: ignore
                  break
              students.add(student)
          for student in students:
            for type in [CORE, ELEC]:
              if self.course_types[type] not in student.takes:
                if not course.overload(student, self.course_types[type]):
                  raise Exception('Impossible')
                break
          
  def run(self):
    self.open_sections_to_overload(self.math)
    self.open_sections_spread_out(self.res)
    for research_group in self.research_groups:
      students     = research_group.students
      demand       = self.get_course_demand(students)
      combinations = self.get_section_combinations(students, demand)
      self.section_grouped(research_group.course, students, combinations)
    for student in self.nogroup_students:
      if not self.enroll_initial(student):
        random.choice(list(student.grade_level.courses[(
          self.res, False)])).overload(student, self.res)
        student.rankings.current(self.math).overload(student, self.math)
    for course in self.courses:
      for section in list(course.sections):
        if len(section.students) < section.capacity.minimum:
          for student in list(section.students):
            for type, course_ in list(student.takes.items()):
              if course == course_:
                student.sections.pop(student.takes.pop(
                  type)).students.remove(student)
                student.sessions.remove(
                  section.parallel_session.session)  # type: ignore
          course.sections.remove(section)
    while True:
      demand = defaultdict[Course, set[tuple[CourseType, Student]]](set)
      for student in self.nogroup_students:
        for course_type in [self.core, self.elec]:
          if course_type not in student.takes:
            course = student.rankings.current(course_type)
            if not course.overload(student, course_type):
              demand[course].add((course_type, student))
      if not self.open_sections_based_on(demand):
        break
    for student in self.students:
      if len(student.takes) != len(self.course_types):
        for course_type in [self.core, self.elec]:
          if course_type in student.takes:
            section = student.sections.pop(student.takes.pop(course_type))
            
            assert section.parallel_session.session
            student.sessions.remove(section.parallel_session.session)
            section.students.remove(student)
        if not self.enroll_final(student):
          raise Exception('Impossible')
    self.rebalance_sections()
    if any(
      len(s.students) < s.capacity.minimum
      for course in self.courses
      for s in course.sections if s.students):
      raise Exception('Some sections are too underloaded')
    
class SolutionV2:
  shifts      : list[Shift]
  courses     : list[Course]
  course_types: dict[str, CourseType]
  
  __research_groups : list[ResearchGroup]
  __students        : list[Student]
  __grouped_students: list[Student]
  __nogroup_students: list[Student]
  
  core: CourseType
  elec: CourseType
  math: CourseType
  res : CourseType
  
  def __init__(self, data: Data):
    self.course_types = data.course_types
    self.shifts       = sorted(data.shifts)
    self.courses      = sorted(data.courses.values())
    
    self.core = self.core
    self.elec = self.elec
    self.math = self.course_types[MATH]
    self.res  = self.course_types[RESEARCH]
    
    self.__research_groups  = list()
    self.__students         = list()
    self.__grouped_students = list()
    self.__nogroup_students = list()
    for student in data.students[data.grade_levels['Grade 12']]:
      if not student.has_taken_level_two:
        core = student.rankings.initial(self.core)
        if not core or core.difficulty_level != 2:
          while student.rankings.current(self.elec).difficulty_level != 2:
            student.rankings.final.pop(
              self.elec, 'Needs to take a level 2 course', 0)
        if not student.research_group:
          research = random.choice(list(
            student.grade_level.courses[(self.res, False)]))
          student.research_group = ResearchGroup(research, 'Temporary')
          student.research_group.add(student)
          
    research_groups  = set[ResearchGroup]()
    grouped_students = set[Student]()
    nogroup_students = set[Student]()
    for students in data.students.values():
      self.__students.extend(students)
      for student in students:
        if student.research_group:
          research_groups.add(student.research_group)
          grouped_students.add(student)
        else:
          nogroup_students.add(student)
    self.__students.sort()
    self.__research_groups  = sorted(research_groups)
    self.__grouped_students = sorted(grouped_students)
    self.__nogroup_students = sorted(nogroup_students)
    
  @property
  def students(self):
    random.shuffle(self.__students)
    return self.__students
  
  @property
  def research_groups(self):
    random.shuffle(self.__research_groups)
    return self.__research_groups
  
  @property
  def grouped_students(self):
    random.shuffle(self.__grouped_students)
    return self.__grouped_students
  
  @property
  def nogroup_students(self):
    random.shuffle(self.__nogroup_students)
    return self.__nogroup_students
          
  def run(self):
    ...
    
def as_text(value: Any):
  if value is None:
    return None
  if any([
    isinstance(value, float),
    isinstance(value, int),
    isinstance(value, str) and value.isdigit()]):
    return int(value)
  return str(value)
    
def read_xlsx(path: str, sheet_name: str):
  workbook = load_workbook(path)
  worksheet = workbook[sheet_name]
  data = list[list[Any]]()
  if isinstance(worksheet, Worksheet):
    for row in worksheet.rows:
      data.append(list(as_text(cell.value) for cell in row))
  workbook.close()
  return data

def encode(data: Data, system_path: str, students_path: str, new: bool):
  if new:
    sheet = read_xlsx(system_path, 'Shifts')
    for r in range(1, len(sheet)):
      shift = Shift()
      for c in range(1, 1 + sheet[r][0]):
        shift.add(Session(sheet[r][c]))
      data.shifts.add(shift)
    
    sheet = read_xlsx(system_path, 'Grade levels')
    for r in range(len(sheet)):
      grade_level = GradeLevel(sheet[r][0])
      data.grade_levels[str(grade_level)] = grade_level
      
    sheet = read_xlsx(system_path, 'Course names')
    for r in range(1, len(sheet)):
      course = Course(sheet[r][0], sheet[r][1])
      data.courses[str(course)] = course
      
    sheet = read_xlsx(system_path, 'Course capacities')
    for r in range(1, len(sheet)):
      course = data.courses[sheet[r][0]]
      course.capacity_section = Capacity(
        sheet[r][1], sheet[r][2], sheet[r][3])
      course.capacity_sections = Capacity(0, 0, sheet[r][4])
    
    sheet = read_xlsx(system_path, 'Course links')
    for r in range(1, len(sheet)):
      data.courses[sheet[r][0]].linked_to = data.courses[sheet[r][1]]
    
    sheet = read_xlsx(system_path, 'Course classification')
    for r in range(1, len(sheet)):
      for c in range(1, len(sheet[r]), 3):
        if any(cell == 'Y' for cell in sheet[r][c:c + 3]):
          course_type = sheet[0][c + 2]
          if course_type not in data.course_types:
            data.course_types[course_type] = CourseType(
              course_type, len(data.course_types))
          data.grade_levels[sheet[0][c]].add(
            data.course_types[course_type],
            sheet[0][c + 1] == 'Y',
            data.courses[sheet[r][0]])
    
    sheet = read_xlsx(system_path, 'Course prerequisites')
    for r in range(1, len(sheet)):
      for c in range(2, 2 + sheet[r][1]):
        data.courses[sheet[r][0]].prerequisites.append(set(
          data.courses[course] for course in sheet[r][c].split('||')))
    
    sheet = read_xlsx(system_path, 'Course not alongside')
    for r in range(1, len(sheet)):
      for c in range(2, 2 + sheet[r][1]):
        data.courses[sheet[r][0]].not_alongside.add(
          data.courses[sheet[r][c]])
    
    sheet = read_xlsx(students_path, 'Research groups')
    for r in range(2, len(sheet)):
      research_group = ResearchGroup(
        data.courses[sheet[0][1]], sheet[r][0])
      data.research_groups[str(research_group)] = research_group
  
  for grade_level_alias in data.grade_levels:
    sheet = read_xlsx(students_path, grade_level_alias)
    for r in range(2, len(sheet)):
      grade_level = data.grade_levels[sheet[r][0]]
      student = Student(sheet[r][1], grade_level)
      
      c = 2
      if sheet[0][c] == 'Groups':
        research_course = data.courses[sheet[1][2]]
        key = f'{research_course} {sheet[r][c]}'
        if key in data.research_groups:
          data.research_groups[key].add(student)
          student.research_group = data.research_groups[key]
        c += 1
      while sheet[0][c] == 'Previous year':
        student.taken.add(data.courses[sheet[r][c]])
        c += 1
      for d in range(c, len(sheet[r])):
        if sheet[r][d]:
          student.rankings.add(
            data.course_types[sheet[1][d]], 
            data.courses[sheet[r][d]])
      data.students[grade_level].append(student)
    data.students[data.grade_levels[grade_level_alias]].sort()
    
def find_filepath(directory: str, template: str):
  path_template = f'{directory}/{template}'
  for index in range(len(next(walk(directory))[2]) + 1):
    if not exists(path_template.format(index)):
      return path_template.format(index)
  return path_template.format(0)

def generate_xlsx(path: str):
  workbook = Workbook(path)
  workbook.close()
  
def delete_default_sheet(path: str):
  workbook = load_workbook(path)
  if 'Sheet1' in workbook.sheetnames:
    worksheet = workbook['Sheet1']
    if not isinstance(worksheet, ReadOnlyWorksheet):
      workbook.remove(worksheet)
  workbook.save(path)
  workbook.close()
  
def realign_columns(path: str, sheet: str):
  workbook = load_workbook(path)
  worksheet = workbook[sheet]
  if isinstance(worksheet, Worksheet):
    for column_cells in worksheet.columns:
      worksheet.column_dimensions[
        get_column_letter(column_cells[0].column)
      ].width = max(len(repr(cell.value)) for cell in column_cells)
  workbook.save(path)
  workbook.close()

def to_xlsx(path: str, sheet: str, data: list[list]):
  if not exists(path):
    generate_xlsx(path)
    
  workbook = load_workbook(path)
  worksheet = workbook.create_sheet(sheet)
  for row in data:
    worksheet.append(as_text(cell) for cell in row)
  workbook.save(path)
  workbook.close()
  
  delete_default_sheet(path)
  realign_columns(path, sheet)
 
def score(data: Data, show_results: Optional[bool] = None):
  total  = defaultdict[CourseType, int](int)
  actual = defaultdict[CourseType, int](int)
  for students in data.students.values():
    for student in students:
      for course_type, ranked in student.grade_level.courses:
        if ranked:
          course = student.rankings.initial(course_type)
          if course:
            total[course_type] += 1
            if course_type in student.takes:
              if course == student.takes[
                course_type] or student.rankings.final.reason_rejected[
                  course_type][course] == 'No rooms available':
                actual[course_type] += 1
  
  scores = dict(
    (course_type, actual[course_type] / total[course_type] * 100)
    for course_type in total)
  if show_results:
    for course_type in scores:
      print(f'{scores[course_type]}% [{course_type}]')
  return scores
    
def meets_target_scores(data: Data, target_score: dict[CourseType, float]):
  scores = score(data, True)
  return not any(list(
    scores[c2] < target_score[c1]
    for c1, c2 in zip(sorted(target_score), sorted(scores))))

def export(
  data: Data, 
  time_taken: int,
  number_of_guesses: int,
  targets: Optional[dict[CourseType, float]] = None):
  filename = find_filepath(RESULT_FILEPATH, RESULT_FILENAME)
  sheet: list[list[Any]] = [
    ['Time taken (s):', time_taken],
    ['Number of guesses:', number_of_guesses]]
  if targets:
    sheet.append(['Course Type', 'Target (%)', 'Score (%)'])
    sheet += list(
      [course_type, targets[course_type], actual_score]
      for course_type, actual_score in score(data).items())
  else:
    sheet.append(['Course Type', 'Score (%)'])
    sheet += list(
      [course_type, actual_score]
      for course_type, actual_score in score(data).items())
  to_xlsx(filename, 'Summary', sheet)
  for grade_level, students in data.students.items():
    course_types = sorted(data.course_types.values())
    ranked_types = sorted(
      course_type for course_type, ranked in grade_level.courses if ranked)
    sheet = [['Student'] + course_types + list(
      map(lambda course_type: f'Remarks: {course_type}', ranked_types))]
    for student in sorted(students):
      line: list[Any] = [student] + list(
        student.sections[student.takes[course_type]] 
        if course_type in student.takes else ''
        for course_type in course_types )
      for course_type in ranked_types:
        if student.rankings.initial(course_type):
          if course_type in student.takes:
            line.append(
              '{}: {}'.format(
                student.rankings.initial(course_type),
                student.rankings.final.reason_rejected[course_type][
                  student.rankings.initial(course_type)
                ])  # type: ignore
              if student.takes[course_type] != student.rankings.initial(
                course_type) else '')
          else:
            line.append(f'Wants {student.rankings.initial(course_type)}')
        else:
          line.append('Initial rankings were invalid')
      sheet.append(line)
    to_xlsx(filename, str(grade_level), sheet)
  for course in data.courses.values():
    if course.sections:
      sheet = list(list() for _ in range(max(map(
        lambda section: len(section.students), course.sections)) + 1))
      for section in sorted(course.sections):
        sheet[0].append(section)
        for i, student in enumerate(sorted(section.students), 1):
          sheet[i].append(student)
        for j in range(i + 1, len(sheet)):  # type: ignore
          sheet[j].append('')
      to_xlsx(filename, str(course), sheet)

def get_target_scores(data: Data):
  course_types = set[CourseType]()
  for grade_level in sorted(data.grade_levels.values()):
    course_types.update(
      course_type for course_type, ranked in grade_level.courses if ranked)
  for course_type in sorted(course_types):
    yield course_type, float(input(f'Target % [{course_type}]: '))
    
def solve(data: Data):
  try:
    SolutionV1(data).run()
  except:
    data.reset()
    return solve(data)
    
def main():
  system_path   = 'input/Test Data_ Subjects.xlsx'
  students_path = 'input/Test Data_ Students.xlsx'
  if not DEBUG:
    system_path   = input('System input file relative path: ')
    students_path = input('Student input file relative path: ')
  if not exists(system_path):
    raise Exception(f'\'{system_path}\' does not exist')
  if not exists(students_path):
    raise Exception(f'\'{students_path} does not exists\'')
  
  data = Data()
  encode(data, system_path, students_path, True)
  
  print('Please select mode:')
  print('  [1] Given target preferences, get a certain number of results')
  print('  [2] Given a certain number of guesses, get the best result')
  match int(input('Mode chosen: ')):
    case 1:
      target_scores = dict(get_target_scores(data))
      results_count = int(input('Number of results to produce: '))
      for _ in range(results_count):
        initial_time = time()
        number_of_guesses = 0
        while not meets_target_scores(data, target_scores):
          number_of_guesses += 1
          solve(data)
        time_taken = ceil(time() - initial_time)
        export(data, time_taken, number_of_guesses, target_scores)
    case 2:
      initial_time = time()
      best = deepcopy(data)
      guess_count = int(input('Number of iterations to choose from: '))
      for _ in range(guess_count):
        solve(data)
        if sum(score(best).values()) < sum(score(data, True).values()):
          best = deepcopy(data)
      time_taken = ceil(time() - initial_time)
      export(best, time_taken, guess_count)
    case _:
      raise Exception('Mode not supported')

if __name__ == '__main__':
  sys.setrecursionlimit(int(1e9))
  
  RESULT_FILEPATH = 'output'
  RESULT_FILENAME = f'{date.today()} Result {"{}"}.xlsx'
  
  RESEARCH = 'Research'
  MATH = 'Mathematics level'
  CORE = 'Core science elective'
  ELEC = 'Science and technology elective'
  
  DEBUG = True
  
  main()