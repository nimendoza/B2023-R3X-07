
    
from __future__  import annotations
from collections import defaultdict
from openpyxl                      import load_workbook
from openpyxl.utils                import get_column_letter
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet  import Worksheet
from os          import walk
from os.path     import exists
from src.classes import *
from typing      import Any, Optional
from xlsxwriter  import Workbook

def as_text(value: Any):
  if value is None:
    return None
  if any([
    isinstance(value, float),
    isinstance(value, int),
    isinstance(value, str) and value.isdigit()]):
    return int(value)
  return str(value)
    
def read_xlsx(path: str, sheet_name: str):
  workbook = load_workbook(path)
  worksheet = workbook[sheet_name]
  data = list[list[Any]]()
  if isinstance(worksheet, Worksheet):
    for row in worksheet.rows:
      data.append(list(as_text(cell.value) for cell in row))
  workbook.close()
  return data

def encode(data: Data, system_path: str, students_path: str):
  sheet = read_xlsx(system_path, 'Shifts')
  for r in range(1, len(sheet)):
    shift = Shift()
    for c in range(1, 1 + sheet[r][0]):
      shift.add(Session(sheet[r][c]))
    data.shifts.add(shift)
  
  sheet = read_xlsx(system_path, 'Grade levels')
  for r in range(len(sheet)):
    grade_level = GradeLevel(sheet[r][0])
    data.grade_levels[str(grade_level)] = grade_level
    
  sheet = read_xlsx(system_path, 'Course names')
  for r in range(1, len(sheet)):
    course = Course(sheet[r][0], sheet[r][1])
    data.courses[str(course)] = course
    
  sheet = read_xlsx(system_path, 'Course capacities')
  for r in range(1, len(sheet)):
    course = data.courses[sheet[r][0]]
    course.capacity_section = Capacity(
      sheet[r][1], sheet[r][2], sheet[r][3])
    course.capacity_sections = Capacity(0, 0, sheet[r][4])
  
  sheet = read_xlsx(system_path, 'Course links')
  for r in range(1, len(sheet)):
    data.courses[sheet[r][0]].linked_to = data.courses[sheet[r][1]]
  
  sheet = read_xlsx(system_path, 'Course classification')
  for r in range(1, len(sheet)):
    for c in range(1, len(sheet[r]), 3):
      if any(cell == 'Y' for cell in sheet[r][c:c + 3]):
        course_type = sheet[0][c + 2]
        if course_type not in data.course_types:
          data.course_types[course_type] = CourseType(
            course_type, len(data.course_types))
        data.grade_levels[sheet[0][c]].add(
          data.course_types[course_type],
          sheet[0][c + 1] == 'Y',
          data.courses[sheet[r][0]])
  
  sheet = read_xlsx(system_path, 'Course prerequisites')
  for r in range(1, len(sheet)):
    for c in range(2, 2 + sheet[r][1]):
      data.courses[sheet[r][0]].prerequisites.append(set(
        data.courses[course] for course in sheet[r][c].split('||')))
  
  sheet = read_xlsx(system_path, 'Course not alongside')
  for r in range(1, len(sheet)):
    for c in range(2, 2 + sheet[r][1]):
      data.courses[sheet[r][0]].not_alongside.add(
        data.courses[sheet[r][c]])
  
  sheet = read_xlsx(students_path, 'Research groups')
  for r in range(2, len(sheet)):
    research_group = ResearchGroup(
      data.courses[sheet[0][1]], sheet[r][0])
    data.research_groups[str(research_group)] = research_group
  
  for grade_level_alias in data.grade_levels:
    sheet = read_xlsx(students_path, grade_level_alias)
    for r in range(2, len(sheet)):
      grade_level = data.grade_levels[sheet[r][0]]
      student = Student(sheet[r][1], grade_level)
      
      c = 2
      if sheet[0][c] == 'Groups':
        research_course = data.courses[sheet[1][2]]
        key = f'{research_course} {sheet[r][c]}'
        if key in data.research_groups:
          data.research_groups[key].add(student)
          student.research_group = data.research_groups[key]
        c += 1
      while sheet[0][c] == 'Previous year':
        student.taken.add(data.courses[sheet[r][c]])
        c += 1
      for d in range(c, len(sheet[r])):
        if sheet[r][d]:
          student.rankings.add(
            data.course_types[sheet[1][d]], 
            data.courses[sheet[r][d]])
      data.students[grade_level].append(student)
    data.students[data.grade_levels[grade_level_alias]].sort()
    
def find_filepath(directory: str, template: str):
  path_template = f'{directory}/{template}'
  for index in range(len(next(walk(directory))[2]) + 1):
    if not exists(path_template.format(index)):
      return path_template.format(index)
  return path_template.format(0)

def generate_xlsx(path: str):
  workbook = Workbook(path)
  workbook.close()
  
def delete_default_sheet(path: str):
  workbook = load_workbook(path)
  if 'Sheet1' in workbook.sheetnames:
    worksheet = workbook['Sheet1']
    if not isinstance(worksheet, ReadOnlyWorksheet):
      workbook.remove(worksheet)
  workbook.save(path)
  workbook.close()
  
def realign_columns(path: str, sheet: str):
  workbook = load_workbook(path)
  worksheet = workbook[sheet]
  if isinstance(worksheet, Worksheet):
    for column_cells in worksheet.columns:
      worksheet.column_dimensions[
        get_column_letter(column_cells[0].column)
      ].width = max(len(repr(cell.value)) for cell in column_cells)
  workbook.save(path)
  workbook.close()

def to_xlsx(path: str, sheet: str, data: list[list]):
  if not exists(path):
    generate_xlsx(path)
    
  workbook = load_workbook(path)
  worksheet = workbook.create_sheet(sheet)
  for row in data:
    worksheet.append(as_text(cell) for cell in row)
  workbook.save(path)
  workbook.close()
  
  delete_default_sheet(path)
  realign_columns(path, sheet)
 
def score(data: Data, show_results: Optional[bool] = None):
  total  = defaultdict[CourseType, int](int)
  actual = defaultdict[CourseType, int](int)
  for students in data.students.values():
    for student in students:
      for course_type, ranked in student.grade_level.courses:
        if ranked:
          course = student.rankings.initial(course_type)
          if course:
            total[course_type] += 1
            if course_type in student.takes:
              if course == student.takes[
                course_type] or student.rankings.final.reason_rejected[
                  course_type][course] == 'No rooms available':
                actual[course_type] += 1
  
  scores = dict(
    (course_type, actual[course_type] / total[course_type] * 100)
    for course_type in total)
  if show_results:
    for course_type in scores:
      print(f'{scores[course_type]}% [{course_type}]')
  return scores
    
def meets_target_scores(data: Data, target_score: dict[CourseType, float]):
  scores = score(data, True)
  return not any(list(
    scores[c2] < target_score[c1]
    for c1, c2 in zip(sorted(target_score), sorted(scores))))
  
def get_target_scores(data: Data):
  course_types = set[CourseType]()
  for grade_level in sorted(data.grade_levels.values()):
    course_types.update(
      course_type for course_type, ranked in grade_level.courses if ranked)
  for course_type in sorted(course_types):
    yield course_type, float(input(f'Target % [{course_type}]: '))
    
def solve(data: Data):
  try:
    SolutionV1(data).run()
  except:
    data.reset()
    return solve(data)