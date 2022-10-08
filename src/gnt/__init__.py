from __future__ import annotations

from collections    import defaultdict
from math           import ceil
from openpyxl       import load_workbook
from openpyxl.utils import get_column_letter
from os             import walk
from os.path        import exists
from src.cls        import Capacity, Category, GradeLevel, Group, ParallelSession, Partition, Section, Shift, Student, Type
from typing         import Any, Iterable
from uuid           import UUID, uuid4
from xlsxwriter     import Workbook

import random

MATH = 'Math'
CORE = 'Core science'
ELEC = 'STE elective'
RESEARCH = 'Research'


class WriteAgent:
    DEFAULT_SHEETNAME = 'Sheet1'

    def as_text(value):
        if value is None:
            return value
        if isinstance(value, float) or isinstance(value, int) or (isinstance(value, str) and value.isdigit()):
            return int(value)
        return str(value)
    
    def generate_xlsx(path: str):
        workbook = Workbook(path)
        workbook.close()

    def delete_default_sheet(path: str):
        workbook = load_workbook(path)
        if WriteAgent.DEFAULT_SHEETNAME in workbook.sheetnames:
            workbook.remove(workbook[WriteAgent.DEFAULT_SHEETNAME])
        workbook.save(path)
        workbook.close()

    def realign_columns(path: str, sheet: str):
        workbook = load_workbook(path)
        worksheet = workbook[sheet]
        for column_cells in worksheet.columns:
            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max(len(repr(cell.value)) for cell in column_cells)
        workbook.save(path)
        workbook.close()
    
    def to_xlsx(path: str, sheet: str, data: list[list]):
        if not exists(path):
            WriteAgent.generate_xlsx(path)

        workbook = load_workbook(path)
        worksheet = workbook.create_sheet(sheet)
        for row in data:
            worksheet.append(WriteAgent.as_text(cell) for cell in row)
        workbook.save(path)
        workbook.close()

        WriteAgent.delete_default_sheet(path)
        WriteAgent.realign_columns(path, sheet)

    def find_filepath(directory: str, name_template: str) -> str:
        path_template = f'{directory}/{name_template}'
        for index in range(len(next(walk(directory))[2]) + 1):
            if not exists(path_template.format(index)):
                return path_template.format(index)

    def get_sheetnames(path: str):
        workbook = load_workbook(path)
        sheetnames = set(workbook.sheetnames)
        workbook.close()
        return sheetnames
    
    def is_sheet_present(path: str, sheet: str):
        return sheet in WriteAgent.get_sheetnames(path)
    
    def find_sheetname(path: str, name_template: str):
        index = 0
        while WriteAgent.is_sheet_present(path, name_template.format(index)):
            index += 1
        return name_template.format(index)
    
    def read_xlsx(path: str, sheet: str):
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook[sheet]
        data = list(list(WriteAgent.as_text(cell.value) for cell in row) for row in worksheet.rows)
        workbook.close()
        return data


class EncodeAgent:
    shifts: set[Shift]
    types: dict[str, Type]
    groups: dict[str, Group]
    categories: dict[str, Category]
    grade_levels: dict[str, GradeLevel]
    students: defaultdict[GradeLevel, list[Student]]

    YES = 'Y'
    GROUP = 'Group'
    DELIMITER = '||'
    SHIFTS = 'Shifts'
    GRADE_LEVEL = 'Grade level'
    PREVIOUS_YEAR = 'Previous year'
    PREREQUISITES = 'Prerequisites'
    NOT_ALONGSIDE = 'Not alongside'
    CLASSIFICATION = 'Classification'
    INFORMATION = 'General information'

    def __init__(self):
        self.shifts = set()
        self.types = dict()
        self.groups = dict()
        self.categories = dict()
        self.grade_levels = dict()
        self.students = defaultdict(list)

    def __repr__(self):
        return 'Encoding Agent'

    def get_subjects(self, path: str):
        # Encode shifts
        data = WriteAgent.read_xlsx(path, EncodeAgent.SHIFTS)
        for r in range(1, len(data)):
            shift = Shift()
            for c in range(1, 1 + data[r][0]):
                shift.add(Partition(data[r][c]))
            self.shifts.add(shift)

        # Encode grade levels
        data = WriteAgent.read_xlsx(path, EncodeAgent.GRADE_LEVEL)
        for r in range(len(data)):
            grade_level = GradeLevel(data[r][0])
            self.grade_levels[str(grade_level)] = grade_level

        # Encode subject information
        ## Encode names and capacities
        data = WriteAgent.read_xlsx(path, EncodeAgent.INFORMATION)
        for r in range(1, len(data)):
            category = Category(data[r][0], data[r][1] if isinstance(data[r][1], int) else None)
            category.capacity_section = Capacity(data[r][2], data[r][4], data[r][3])
            category.capacity_sections = Capacity(0, data[r][5])
            self.categories[str(category)] = category

        ## Encode linked subjects
        for r in range(1, len(data)):
            key = data[r][6]
            if key in self.categories:
                self.categories[str(Category(data[r][0], data[r][1] if isinstance(data[r][1], int) else None))].link = self.categories[key]

        ## Encode classification
        data = WriteAgent.read_xlsx(path, EncodeAgent.CLASSIFICATION)
        for r in range(1, len(data)):
            for c in range(1, len(data[r]), 3):
                if any(cell == EncodeAgent.YES for cell in data[r][c:c + 3]):
                    type  = data[0][c + 2]
                    if type not in self.types:
                        self.types[type] = Type(type, len(self.types))
                    self.grade_levels[data[0][c]].add(self.types[type], data[0][c + 1] == EncodeAgent.YES, self.categories[data[r][0]])

        ## Encode prerequisites
        data = WriteAgent.read_xlsx(path, EncodeAgent.PREREQUISITES)
        data = WriteAgent.read_xlsx(path, self.PREREQUISITES)
        for r in range(1, len(data)):
            for c in range(2, 2 + data[r][1]):
                prerequisites = tuple(self.categories[s] for s in data[r][c].split(self.DELIMITER))
                self.categories[data[r][0]].add_prerequisites(prerequisites)
        
        ## Encode not alongsides
        data = WriteAgent.read_xlsx(path, self.NOT_ALONGSIDE)
        for r in range(1, len(data)):
            for c in range(2, 2 + data[r][1]):
                self.categories[data[r][0]].add_not_alongside(self.categories[data[r][c]])

    def get_students(self, path: str):
        # Encode groups
        for sheet in WriteAgent.get_sheetnames(path):
            if sheet.find(EncodeAgent.GROUP) != -1:
                data = WriteAgent.read_xlsx(path, sheet)
                for r in range(2, len(data)):
                    group = Group(data[r][0], self.categories[data[0][1]])
                    self.groups[str(group)] = group
        
        # Encode students
        for sheet in WriteAgent.get_sheetnames(path):
            if sheet.find(self.GRADE_LEVEL) != -1:
                data = WriteAgent.read_xlsx(path, sheet)
                for r in range(2, len(data)):
                    grade_level = self.grade_levels[data[r][0]]
                    student = Student(data[r][1], grade_level)

                    c = 2
                    if data[0][c].upper().find(EncodeAgent.GROUP.upper()) != -1:
                        ## Encode group
                        parent = self.categories[data[1][2]]
                        key = f'{parent} {data[r][c]}'
                        if key in self.groups:
                            self.groups[key].add(student)
                            student.group = self.groups[key]
                        c += 1
                    while data[0][c].upper() == self.PREVIOUS_YEAR.upper():
                        student.taken.add(self.categories[data[r][c]])
                        c += 1
                    for d in range(c, len(data[r])):
                        if data[r][d]:
                            student.rankings.add(self.types[data[1][d]], self.categories[data[r][d]])
                    self.students[grade_level].append(student)

    def clean(self):
        for students in self.students.values():
            for student in filter(lambda x: x.rankings.current(self.types[CORE]) in x.rankings.final.ordered[self.types[ELEC]], students):
                final_ranks = student.rankings.final
                start_ranks = student.rankings.start
                category = start_ranks.ordered[self.types[ELEC]].index(student.rankings.initial(self.types[CORE]))
                final_ranks.pop(self.types[ELEC], 'Already the first choice Core elective', category)
                start_ranks.pop(self.types[ELEC], 'Already the first choice Core elective', category)
                
                
class AnalyzeAgent:
    def score(encode_agent: EncodeAgent, show: bool = None):
        total = defaultdict[Type, int](int)
        attained = defaultdict[Type, int](int)
        for students in encode_agent.students.values():
            for student in students:
                for type, _ in filter(lambda x: x[1], student.grade_level.categories):
                    total[type] += 1
                    if type in student.takes and (not student.rankings.initial(type) or student.rankings.initial(type) == student.takes[type]):
                        attained[type] += 1
        
        scores = dict((type, attained[type] / total[type] * 100) for type in total)
        if show:
            for type in scores:
                print(f'{scores[type]}% [{type}]')
        return scores
    
    
class AllocateAgent:
    shifts: list[Shift]
    types: dict[str, Type]
    cateogries: list[Category]
    
    __groups: list[Group]
    __students: list[Student]
    __grouped_students: list[Student]
    __nogroup_students: list[Student]

    def __init__(self, encode_agent: EncodeAgent):
        self.types = encode_agent.types
        self.shifts = list(encode_agent.shifts)
        self.categories = list(encode_agent.categories.values())
        
        self.__students = list()

        groups = set[Group]()
        grouped_students = set[Student]()
        nogroup_students = set[Student]()
        for students in encode_agent.students.values():
            self.__students.extend(students)
            for student in students:
                if student.group:
                    groups.add(student.group)
                    grouped_students.add(student)
                else:
                    nogroup_students.add(student)
        self.__groups = list(groups)
        self.__grouped_students = list(grouped_students)
        self.__nogroup_students = list(nogroup_students)

    @property
    def students(self):
        random.shuffle(self.__students)
        return self.__students

    @property
    def groups(self):
        random.shuffle(self.__groups)
        return self.__groups

    @property
    def grouped_students(self):
        random.shuffle(self.__grouped_students)
        return self.__grouped_students

    @property
    def nogroup_students(self):
        random.shuffle(self.__nogroup_students)
        return self.__nogroup_students
    
    def preopen_sections(self, type: Type):
        demand = defaultdict[Category, int](int)
        for student in self.students:
            demand[student.rankings.current(type)] += 1

        to_open = dict((category, ceil(demand[category] / category.capacity_section.maximum)) for category in demand)
        for category in to_open:
            if category.link:
                assert category.capacity_sections == category.link.capacity_sections
                while sum(to_open[c] for c in [category, category.link]) > category.capacity_sections.maximum:
                    minimum = category if demand[category] % category.capacity_section.maximum < demand[category.link] % category.link.capacity_section.maximum else category.link
                    to_open[minimum] -= 1
            else:
                to_open[category] = min(to_open[category], category.capacity_sections.maximum)
        for category, amount in to_open.items():
            for _ in range(amount):
                section = Section(category)
                category.sections.add(section)
                
    def get_session(self, category: Category, shift: Shift, partition: Partition):
        return ParallelSession(shift, partition, len(category.list_sections_by(partition)))
    
    def try_section_student(self, student: Student, type: Type, shift: Shift, partitions: list[Partition]):
        category = student.rankings.current(type)
        if category.qualified(student):
            for partition in partitions:
                for section in category.list_sections_by(partition):
                    if section.enroll(student, type):
                        return True
            return False
        student.rankings.final.pop(type, 'Not qualified')
        return self.try_section_student(student, type, shift, partitions)
    
    def section_student(self, student: Student, type: Type, shift: Shift, partitions: list[Partition]):
        category = student.rankings.current(type)
        if category.qualified(student):
            for partition in partitions:
                for section in category.list_sections_by(partition):
                    if section.enroll(student, type):
                        return
            if category.could_open_section():
                section = Section(category, shift, self.get_session(category, shift, random.choice(partitions)))
                category.sections.add(section); assert section.enroll(student, type)
                return
            student.rankings.final.pop(type, 'No rooms available')
        else:
            student.rankings.final.pop(type, 'Not qualified')
        self.section_student(student, type, shift, partitions)
    
    def section_grouped(self):
        for group in self.groups:
            if not group.parent.sections:
                for shift in self.shifts:
                    for partition in shift.partitions:
                        group.parent.sections.add(Section(group.parent, shift, self.get_session(group.parent, shift, partition)))
            demand = defaultdict[Category, int](int)
            for student in group.students:
                core = student.rankings.current(self.types[CORE])
                elec = student.rankings.current(self.types[ELEC])
                while elec in core.not_alongside or elec == core:
                    student.rankings.final.pop(self.types[ELEC], 'Could not be taken alongside Core science')
                    elec = student.rankings.current(self.types[ELEC])
                demand[core] += 1
                demand[elec] += 1
            section_combinations = defaultdict[Shift, set[tuple[Student, Section, Section]]](set)
            for student in group.students:
                core = student.rankings.current(self.types[CORE])
                elec = student.rankings.current(self.types[ELEC])
                for capacity_section in filter(lambda x: len(x.students) + demand[x.parent] <= x.capacity.maximum, core.sections):
                    for esection in filter(lambda x: len(x.students) + demand[x.parent] <= x.capacity.maximum, elec.sections):
                        if capacity_section.shift == esection.shift and capacity_section.session.partition != esection.session.partition:
                            section_combinations[capacity_section.shift].add((student, capacity_section, esection))
            section_excluded_partitions = defaultdict[tuple[Shift, Partition], set[tuple[Student, Section, Section]]](set)
            for shift in section_combinations:
                for partition in shift.partitions:
                    for student, capacity_section, esection in section_combinations[shift]:
                        if capacity_section.session.partition != partition and esection.session.partition != partition:
                            section_excluded_partitions[(shift, partition)].add((student, capacity_section, esection))
            research_sections = list(filter(lambda x: len(x.students) + len(group.students) <= x.capacity.maximum, group.parent.sections))
            if section_excluded_partitions:
                for research_section in research_sections:
                    if (research_section.shift, research_section.session.partition) not in section_excluded_partitions:
                        continue
                    for student in group.students:
                        research_section.enroll(student, self.types[RESEARCH])
                        if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                            student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
                    for student, capacity_section, esection in section_excluded_partitions[(research_section.shift, research_section.session.partition)]:
                        if not set(student.takes).intersection({self.types[CORE], self.types[ELEC]}):
                            capacity_section.enroll(student, self.types[CORE])
                            esection.enroll(student, self.types[ELEC])
                    for student in filter(lambda x: not set(x.takes).intersection({self.types[CORE], self.types[ELEC]}), group.students):
                        if not self.try_section_student(student, self.types[CORE], student.shift, student.available_partitions):
                            self.section_student(student, self.types[CORE], student.shift, student.available_partitions)
                        if not self.try_section_student(student, self.types[ELEC], student.shift, student.available_partitions):
                            self.section_student(student, self.types[ELEC], student.shift, student.available_partitions)
                    break
                if not group.shift:
                    assert group.parent.could_open_section()
                    shift, partition = max(section_excluded_partitions, key=lambda x: len(section_excluded_partitions[x]))
                    research_section = Section(group.parent, shift, self.get_session(group.parent, shift, partition))
                    group.parent.sections.add(research_section)
                    for student in group.students:
                        research_section.enroll(student, self.types[RESEARCH])
                        if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                            student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
                    for student, capacity_section, esection in section_excluded_partitions[(shift, partition)]:
                        if not set(student.takes).intersection({self.types[CORE], self.types[ELEC]}):
                            capacity_section.enroll(student, self.types[CORE])
                            esection.enroll(student, self.types[ELEC])
                    for student in filter(lambda x: not set(x.takes).intersection({self.types[CORE], self.types[ELEC]}), group.students):
                        if not self.try_section_student(student, self.types[CORE], student.shift, student.available_partitions):
                            self.section_student(student, self.types[CORE], student.shift, student.available_partitions)
                        if not self.try_section_student(student, self.types[ELEC], student.shift, student.available_partitions):
                            self.section_student(student, self.types[ELEC], student.shift, student.available_partitions)
            else:
                if not research_sections and group.parent.could_open_section():
                    shift = random.choice(self.shifts)
                    research_section = Section(group.parent, shift, self.get_session(group.parent, shift, random.choice(list(shift.partitions))))
                    group.parent.sections.add(research_section)
                    research_sections.append(research_section)
                research_section = random.choice(research_sections)
                for student in group.students:
                    research_section.enroll(student, self.types[RESEARCH])
                    if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                        student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
                    if not self.try_section_student(student, self.types[CORE], student.shift, student.available_partitions):
                        self.section_student(student, self.types[CORE], student.shift, student.available_partitions)
                    if not self.try_section_student(student, self.types[ELEC], student.shift, student.available_partitions):
                        self.section_student(student, self.types[ELEC], student.shift, student.available_partitions)
                        
    def section_nogroup(self):
        def enroll(student: Student):
            research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))            
            core = student.rankings.current(self.types[CORE])
            elec = student.rankings.current(self.types[ELEC])
            while elec == core or elec in core.not_alongside:
                student.rankings.final.pop(self.types[ELEC], 'Could not be alongside Core science')
                elec = student.rankings.current(self.types[ELEC])
            for shift in self.shifts:
                for c in core.list_sections_by(shift):
                    for e in elec.list_sections_by(shift):
                        for r in research.list_sections_by(shift):
                            if all([
                                c.session.partition != e.session.partition,
                                c.session.partition != r.session.partition,
                                e.session.partition != r.session.partition,
                                len(c.students) < c.capacity.maximum,
                                len(e.students) < e.capacity.maximum,
                                len(r.students) < r.capacity.maximum
                            ]):
                                assert c.enroll(student, self.types[CORE]) and e.enroll(student, self.types[ELEC]) and r.enroll(student, self.types[RESEARCH])
                                if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                                    student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
                                return True
                            
        for student in self.nogroup_students:
            research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))  
            if not research.sections:
                for shift in self.shifts:
                    for partition in shift.partitions:
                        research.sections.add(Section(research, shift, self.get_session(research, shift, partition)))
            if not enroll(student):
                for type in [self.types[CORE], self.types[ELEC]]:
                    category = student.rankings.current(type)
                    category.enroll(student, type)
                if student.shift:
                    if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                        student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
                    research.enroll(student, self.types[RESEARCH])
            continue
            core = student.rankings.current(self.types[CORE])
            elec = student.rankings.current(self.types[ELEC])
            core.enroll(student, self.types[CORE])
            elec.enroll(student, self.types[ELEC])
            if student.shift:
                if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                    student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
        for student in self.nogroup_students:
            research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))
            for research_section in list(research.sections):
                if not research_section.students:
                    research.sections.remove(research_section)
        return
        for _ in range(10):
            shifted_demand = dict[Shift, defaultdict[Category, set[tuple[Type, Student]]]]((shift, defaultdict(set)) for shift in self.shifts)
            noshift_demand = defaultdict[Category, set[tuple[Type, Student]]](set)
            for student in self.nogroup_students:
                for type in filter(lambda x: x not in student.takes, [self.types[CORE], self.types[ELEC]]):
                    if student.shift:
                        shifted_demand[student.shift][student.rankings.current(type)].add((type, student))
                    else:
                        noshift_demand[student.rankings.current(type)].add((type, student))
            for shift, demand in shifted_demand.items():
                for category, pairs in sorted(demand.items(), key=lambda x: len(x[1]), reverse=True):
                    partitioned = defaultdict[Partition, set[tuple[Type, Student]]](set)
                    for type, student in pairs:
                        for partition in student.available_partitions:
                            partitioned[partition].add((type, student))
                    for partition, pairs_ in sorted(partitioned.items(), key=lambda x: len(x[1]), reverse=True):
                        if len(pairs_) + len(noshift_demand[category]) >= category.capacity_section.minimum and category.could_open_section():
                            category.sections.add(Section(category, shift, self.get_session(category, shift, partition)))
                            for type, student in pairs:
                                category.enroll(student, type)
                            for type, student in list(noshift_demand[category]):
                                if category.enroll(student, type):
                                    noshift_demand[category].remove((type, student))
                            break
        for student in self.nogroup_students:
            if self.types[CORE] in student.takes and self.types[ELEC] in student.takes and self.types[RESEARCH] not in student.takes:
                research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))  
                research.enroll(student, self.types[RESEARCH])
            if student.shift and self.types[MATH] not in student.takes:
                if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                    student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
        for student in self.students:
            for type in filter(lambda x: x not in student.takes, [self.types[CORE], self.types[ELEC]]):
                for category in student.grade_level.categories[(type, True)]:
                    if category.enroll(student, type):
                        if category != student.rankings.initial(type) and student.rankings.initial(type):
                            student.rankings.final.pop(type, 'No more rooms available')
                        break
            if self.types[RESEARCH] not in student.takes:
                research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))  
                research.enroll(student, self.types[RESEARCH])
            if student.shift and self.types[MATH] not in student.takes:
                if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                    student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
        for student in self.nogroup_students:
            research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))
            for research_section in list(research.sections):
                if not research_section.students:
                    research.sections.remove(research_section)
        
        research_demand = dict[Category, defaultdict[tuple[Shift, Partition], set[tuple[Type, Student]]]]((random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)])), defaultdict(set)) for student in self.students)
        for student in self.students:
            research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))  
            if self.types[RESEARCH] not in student.takes and not research.enroll(student, self.types[RESEARCH]):
                # assert len(student.available_partitions) == 1
                research_demand[research][(student.shift, random.choice(student.available_partitions))].add((self.types[RESEARCH], student))
        for category, demand in research_demand.items():
            for pair, pairs in sorted(demand.items(), key=lambda x: len(x[1]), reverse=True):
                if category.could_open_section():
                    category.sections.add(Section(category, pair[0], self.get_session(category, pair[0], pair[1])))
                    for type, student in pairs:
                        category.enroll(student, type)
        for student in self.students:
            if self.types[RESEARCH] not in student.takes:
                for type, category in list(student.takes.items()):
                    student.takes.pop(type)
                    student.sections.pop(category).students.remove(student)
                student.shift = None
                student.sessions = set()
                research = random.choice(list(student.grade_level.categories[(self.types[RESEARCH], False)]))  
                if not enroll(student):
                    if not research.enroll(student, self.types[RESEARCH]):
                        research.overload(student, self.types[RESEARCH])
                for type in filter(lambda x: x not in student.takes, [self.types[CORE], self.types[ELEC]]):
                    for category in student.grade_level.categories[(type, True)]:
                        if category.enroll(student, type):
                            if category != student.rankings.initial(type) and student.rankings.initial(type):
                                student.rankings.final.pop(type, 'No more rooms available')
                            break
                if student.shift and self.types[MATH] not in student.takes:
                    if not student.rankings.current(self.types[MATH]).enroll(student, self.types[MATH]):
                        student.rankings.current(self.types[MATH]).overload(student, self.types[MATH])
        for _ in range(10):
            shifted_demand = dict[Shift, defaultdict[Category, set[tuple[Type, Student]]]]((shift, defaultdict(set)) for shift in self.shifts)
            noshift_demand = defaultdict[Category, set[tuple[Type, Student]]](set)
            for student in self.nogroup_students:
                for type in filter(lambda x: x not in student.takes, [self.types[CORE], self.types[ELEC]]):
                    if student.shift:
                        shifted_demand[student.shift][student.rankings.current(type)].add((type, student))
                    else:
                        noshift_demand[student.rankings.current(type)].add((type, student))
            for shift, demand in shifted_demand.items():
                for category, pairs in sorted(demand.items(), key=lambda x: len(x[1]), reverse=True):
                    partitioned = defaultdict[Partition, set[tuple[Type, Student]]](set)
                    for type, student in pairs:
                        for partition in student.available_partitions:
                            partitioned[partition].add((type, student))
                    for partition, pairs_ in sorted(partitioned.items(), key=lambda x: len(x[1]), reverse=True):
                        if len(pairs_) + len(noshift_demand[category]) >= category.capacity_section.minimum and category.could_open_section():
                            category.sections.add(Section(category, shift, self.get_session(category, shift, partition)))
                            for type, student in pairs:
                                category.enroll(student, type)
                            for type, student in list(noshift_demand[category]):
                                if category.enroll(student, type):
                                    noshift_demand[category].remove((type, student))
                            break
        for student in self.students:
            for type in filter(lambda x: x not in student.takes, [self.types[CORE], self.types[ELEC]]):
                for category in student.grade_level.categories[(type, True)]:
                    if category.enroll(student, type):
                        if category != student.rankings.initial(type) and student.rankings.initial(type):
                            student.rankings.final.pop(type, 'No more rooms available')
                        break
            assert len(student.takes) == 4
    
    def solve(self):
        self.preopen_sections(self.types[MATH])
        self.section_grouped()
        self.section_nogroup()
        # assert all(len(s.students) >= s.capacity.minimum for category in self.categories for s in category.sections if s.students)
        
        
def run():
    encode_agent = EncodeAgent()
    encode_agent.get_subjects('input/Test Data_ Subjects.xlsx')
    encode_agent.get_students('input/Test Data_ Students.xlsx')
    encode_agent.clean()

    allocate_agent = AllocateAgent(encode_agent)
    allocate_agent.solve()
    return encode_agent

def export(encode_agent: EncodeAgent, directory: str, template: str):
    filename = WriteAgent.find_filepath(directory, template)
    data = [['Course Type', 'Score (%)']]
    for type, score in AnalyzeAgent.score(encode_agent).items():
        data.extend([[type, score]])
    WriteAgent.to_xlsx(filename, 'Summary', data)
    for grade_level, students in encode_agent.students.items():
        types = sorted(encode_agent.types.values())
        data = [['Student']]
        data[0].extend(types)
        data[0].extend(map(lambda x: f'Remarks: {x}', types[:3]))
        for student in students:
            line = [student]
            for type in types:
                line.append(student.sections[student.takes[type]] if type in student.takes else '')
            for type in types[:3]:
                if student.rankings.initial(type):
                    if type in student.takes:
                        line.append(f'{student.rankings.initial(type)}: {student.rankings.final.reasons[type][student.rankings.initial(type)]}' if student.takes[type] != student.rankings.initial(type) else '')
                    else:
                        line.append(f'Wants {student.rankings.initial(type)}')
                else:
                    line.append('Initial rankings are invalid')
            data.append(line)
        WriteAgent.to_xlsx(filename, str(grade_level), data)
    for category in filter(lambda x: x.sections, encode_agent.categories.values()):
        data = list(list() for _ in range(max(map(lambda x: len(x.students), category.sections)) + 1))
        for section in category.sections:
            data[0].append(section)
            for i, student in enumerate(section.students, 1):
                data[i].append(student)
        WriteAgent.to_xlsx(filename, str(category), data)

def main():
    encode_agent = None
    while not encode_agent:
        try:
            encode_agent = run()
        except:
                continue
        break
        if min(AnalyzeAgent.score(encode_agent, True).values()) < 90:
            encode_agent = None
            continue
        export(encode_agent, 'output', 'Test Results {}.xlsx')
    AnalyzeAgent.score(encode_agent, True)
    export(encode_agent, 'output', 'Test Results {}.xlsx')

if __name__ == '__main__':
    main()